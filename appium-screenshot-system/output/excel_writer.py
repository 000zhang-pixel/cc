from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import TYPE_CHECKING

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from loguru import logger

from utils.helpers import ensure_dir, safe_filename

if TYPE_CHECKING:
    from models.content_item import ContentItem, SearchSession


# Colour palette
COLOUR = {
    "header_bg": "2D6A4F",
    "header_fg": "FFFFFF",
    "pos": "D8F3DC",    # positive sentiment
    "neg": "FFCCD5",    # negative sentiment
    "neutral": "F0F0F0",
    "high_risk": "FF4D4D",
    "row_alt": "F7F7F7",
}


class ExcelWriter:

    def __init__(self, output_dir: str = "output/data", timestamp: bool = True):
        self.output_dir = ensure_dir(output_dir)
        self.timestamp = timestamp

    # ──────────────────────────── public API ────────────────────────────

    def write_session(self, session: "SearchSession",
                      analysis: dict | None = None) -> Path:
        """为单次搜索会话生成 Excel 文件"""
        ts = f"_{datetime.now().strftime('%Y%m%d_%H%M%S')}" if self.timestamp else ""
        fname = f"{safe_filename(session.platform)}_{safe_filename(session.keyword)}{ts}.xlsx"
        path = self.output_dir / fname

        with pd.ExcelWriter(path, engine="openpyxl") as writer:
            # Sheet1: 原始数据
            self._write_raw(writer, session.items)
            # Sheet2: 分析摘要
            if analysis:
                self._write_analysis(writer, analysis)

        self._apply_styles(path, len(session.items), analysis)
        logger.info(f"Excel 已保存: {path}")
        return path

    def write_multi_sessions(self, sessions: list["SearchSession"],
                              analyses: list[dict] | None = None) -> Path:
        """将多个会话合并到一个 Excel，每个会话一个 sheet"""
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        path = self.output_dir / f"全量报告_{ts}.xlsx"

        with pd.ExcelWriter(path, engine="openpyxl") as writer:
            all_items: list["ContentItem"] = []
            for s in sessions:
                all_items.extend(s.items)
                if s.items:
                    sheet_name = f"{s.platform[:4]}_{s.keyword[:8]}"
                    df = pd.DataFrame([i.to_dict() for i in s.items])
                    df.to_excel(writer, sheet_name=sheet_name, index=False)

            # 汇总 sheet
            if all_items:
                df_all = pd.DataFrame([i.to_dict() for i in all_items])
                df_all.to_excel(writer, sheet_name="全量数据", index=False)

            # 分析汇总 sheet
            if analyses:
                summary_rows = []
                for a in analyses:
                    from analysis.ranking import RankingAnalyzer
                    summary_rows.extend(RankingAnalyzer().summary_table(a))
                df_summary = pd.DataFrame(summary_rows)
                df_summary.to_excel(writer, sheet_name="舆情分析汇总", index=False)

        self._style_multi(path)
        logger.info(f"全量 Excel 已保存: {path}")
        return path

    # ──────────────────────────── private helpers ────────────────────────

    def _write_raw(self, writer, items: list["ContentItem"]) -> None:
        rows = [i.to_dict() for i in items]
        df = pd.DataFrame(rows) if rows else pd.DataFrame()
        df.to_excel(writer, sheet_name="搜索结果", index=False)

    def _write_analysis(self, writer, analysis: dict) -> None:
        from analysis.ranking import RankingAnalyzer
        rows = RankingAnalyzer().summary_table(analysis)
        df = pd.DataFrame(rows)
        df.to_excel(writer, sheet_name="舆情分析", index=False)

    # ──────────────────────────── styling ────────────────────────────────

    def _apply_styles(self, path: Path, data_rows: int,
                      analysis: dict | None) -> None:
        wb = load_workbook(path)
        for ws in wb.worksheets:
            self._style_sheet(ws)
        wb.save(path)

    def _style_multi(self, path: Path) -> None:
        wb = load_workbook(path)
        for ws in wb.worksheets:
            self._style_sheet(ws)
        wb.save(path)

    def _style_sheet(self, ws) -> None:
        if ws.max_row < 1:
            return

        header_fill = PatternFill("solid", fgColor=COLOUR["header_bg"])
        header_font = Font(bold=True, color=COLOUR["header_fg"], size=11)
        border = Border(
            left=Side(style="thin", color="CCCCCC"),
            right=Side(style="thin", color="CCCCCC"),
            top=Side(style="thin", color="CCCCCC"),
            bottom=Side(style="thin", color="CCCCCC"),
        )

        # Header row
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border

        # Data rows
        sentiment_col = self._col_index(ws, "情感倾向")
        risk_col = self._col_index(ws, "风险等级")

        for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            fill_color = COLOUR["row_alt"] if row_idx % 2 == 0 else "FFFFFF"

            for cell in row:
                cell.alignment = Alignment(vertical="center", wrap_text=False)
                cell.border = border
                # Conditional colouring
                if cell.column == sentiment_col and cell.value:
                    fill_color_cell = (
                        COLOUR["pos"] if cell.value == "positive" else
                        COLOUR["neg"] if cell.value == "negative" else
                        COLOUR["neutral"]
                    )
                    cell.fill = PatternFill("solid", fgColor=fill_color_cell)
                elif cell.column == risk_col and cell.value == "high":
                    cell.fill = PatternFill("solid", fgColor=COLOUR["high_risk"])
                    cell.font = Font(bold=True, color="FFFFFF")
                else:
                    cell.fill = PatternFill("solid", fgColor=fill_color)

        # Auto-fit column width
        for col in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col), default=0)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 40)

        ws.freeze_panes = "A2"

    @staticmethod
    def _col_index(ws, header: str) -> int | None:
        for cell in ws[1]:
            if cell.value == header:
                return cell.column
        return None
